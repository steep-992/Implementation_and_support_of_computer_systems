import random
import openpyxl
from openpyxl.styles import Font, Alignment


def generate_adjutanted():
    fio = input("Введите ФИО: ")
    month = input("Введите месяц и год (например, Дек. 2024): ")
    hours_worked = int(input("Введите количество отработанных часов: "))
    filename = "adjutanted.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Расчетный лист"

    ws.merge_cells("A1:G1")
    ws["A1"] = f"РАСЧЕТНЫЙ ЛИСТОК {month}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws["A2"] = "ФИО:"
    ws["B2"] = fio
    ws["A3"] = "Отработанные часы:"
    ws["B3"] = hours_worked

    headers = ["Вид", "Период", "Рабочие дни", "Часы", "Оплачено", "Сумма"]
    ws.append(headers)

    for col in range(1, len(headers) + 1):
        ws.cell(row=5, column=col).font = Font(bold=True)


    payroll_data = [
        ("Должностной оклад", month, random.randint(15, 22), hours_worked, hours_worked, hours_worked * 80),
        ("Доплата за объем работ", month, "-", hours_worked * 0.1, hours_worked * 0.1, hours_worked * 8),
        ("Персональная надбавка", month, "-", hours_worked * 0.2, hours_worked * 0.2, hours_worked * 12),
        ("Районный коэффициент", month, "-", hours_worked * 0.15, hours_worked * 0.15, hours_worked * 10),
        ("Северная надбавка", month, "-", hours_worked * 0.1, hours_worked * 0.1, hours_worked * 9),
        ("Надбавка за стаж", month, "-", hours_worked * 0.05, hours_worked * 0.05, hours_worked * 6),
    ]

    for row in payroll_data:
        ws.append(row)

    total_sum = sum(row[5] for row in payroll_data)
    ws.append(["Итого", "", "", "", "", total_sum])
    ws.cell(row=ws.max_row, column=6).font = Font(bold=True)

    wb.save(filename)
    print(f"Файл '{filename}' успешно создан!")

generate_adjutanted()